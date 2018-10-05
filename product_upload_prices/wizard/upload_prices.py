# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from openerp import api, models, fields, _
from openerp.exceptions import UserError
import base64
import tempfile
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except (ImportError, IOError) as err:
    _logger.debug(err)


class UploadPrices(models.TransientModel):
    _name = "product_upload_prices.upload_prices"

    data = fields.Binary(
        'File',
        required=True
    )
    name = fields.Char(
        'Filename',
        readonly=True
    )

    @staticmethod
    def read_data(sheet):
        """ Read the spreadsheet into a data structure
        """
        ret = []
        row_number = 0
        for row in sheet.iter_rows(min_row=1, min_col=1,
                                   max_col=3, max_row=40000):
            row_number += 1
            ret.append({
                'default_code': str(row[0].value),
                'list_price': row[1].value,
                'standard_price': row[2].value,
                'row': row_number
            })
        return ret

    def check_data(self, data, vendor):
        """ Check data structure for errors
        """
        product_obj = self.env['product.product']
        for row in data:
            # check product exists
            domain = [('default_code', '=', row['default_code'])]
            if not product_obj.search(domain):
                raise UserError(
                    _(u'ERROR in line %s, from vendor %s, product '
                      u'"%s" not found') %
                    (row['row'], vendor, row['default_code']))
            try:
                # check list price is a number
                float(row['list_price'])
            except (ValueError, TypeError):
                raise UserError(
                    _(u'Error in line %s, from vendor %s list price "%s" '
                      u'is not a number') %
                    (row['row'], vendor, row['list_price']))
            try:
                # check standard price is a number
                float(row['standard_price'])
            except (ValueError, TypeError):
                raise UserError(
                    _(u'Error in line %s, from vendor %s, standard price '
                      u'"%s" is not a number') %
                    (row['row'], vendor, row['standard_price']))

    def process_data(self, data, vendor):
        """ Process data structure setting prices in system
        """
        product_obj = self.env['product.template']
        for row in data:
            domain = [('default_code', '=', row['default_code'])]
            prod = product_obj.search(domain)

            price = float(row['list_price'])
            cost = float(row['standard_price'])

            prod.set_prices(cost, vendor, price=price)
            prod.set_invoice_cost()
            _logger.info('Importing price %s' % prod.default_code)

    @api.multi
    def import_file(self):
        data = base64.decodestring(self.data)
        (fileno, fp_name) = tempfile.mkstemp('.xlsx', 'openerp_')

        # escribir la planilla en un temporario
        openfile = open(fp_name, "w")
        openfile.write(data)
        openfile.close()

        # leer la planilla del temporario en solo lectura
        wb = openpyxl.load_workbook(filename=fp_name, read_only=True,
                                    data_only=True)

        partner_obj = self.env['res.partner']
        # cada hoja de la planilla es un vendor
        vendors = wb.sheetnames
        for vendor in vendors:
            partner = partner_obj.search([('ref', '=', vendor)])
            if not partner:
                raise UserError(_('Vendor %s not found.') % vendor)
            data = self.read_data(wb[vendor])
            self.check_data(data, vendor)
            self.process_data(data, vendor)
