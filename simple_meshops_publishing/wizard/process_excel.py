# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from openerp import api, models, fields, _
import base64
import tempfile
import openpyxl
import os


class SimpleMeShopsPublishing(models.TransientModel):
    _name = "simple_meshops_publishing.process_excel"
    data = fields.Binary(
        string="Worksheet to process"
    )
    name = fields.Char(
        'File Name',
    )
    state = fields.Selection(
        [('process', 'Process'),  # load spreadsheet
         ('download', 'Download')],  # download spreadsheet
        default="process"
    )
    date_from = fields.Date(
        default=fields.date.today()
    )

    @api.one
    def process_data(self, fp_name):
        CODE_COL = 1
        PRICE_COL = 2
        FIRST_ROW = 2

        product_obj = self.env['product.product']
        products = product_obj.search([('meshops_code', '=', True),
                                       ('write_date', '>=', self.date_from)])

        # open worksheet
        wb = openpyxl.load_workbook(filename=fp_name,
                                    read_only=False,
                                    data_only=True)
        sheet = wb.active
        row = FIRST_ROW
        for product in products:
            sheet.cell(column=CODE_COL,
                       row=row).value = product.default_code
            sheet.cell(column=PRICE_COL,
                       row=row).value = product.final_price
            row += 1
        wb.save(fp_name)

    @api.multi
    def process_spreadsheet(self):
        # mover la planilla a un temporario
        (fileno, fp_name) = tempfile.mkstemp('.xlsx', 'openerp_')
        file = os.path.dirname(os.path.realpath(__file__))
        file = file.replace('wizard', 'data/meshop_prices.xlsx')

        with open(file, "r") as worksheet:
            data = worksheet.read()
        with open(fp_name, "w") as worksheet:
            worksheet.write(data)

        # procesar la planilla
        self.process_data(fp_name)

        # preparar para download
        with open(fp_name, "r") as worksheet:
            data = worksheet.read()

        for rec in self:
            rec.data = base64.encodestring(data)
            rec.state = 'download'
            rec.name = 'planilla.xlsx'
