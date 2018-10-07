# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from openerp import api, models, fields, _
import base64
import tempfile
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except (ImportError, IOError) as err:
    _logger.debug(err)


class SimpleMeliPublishing(models.TransientModel):
    _name = "simple_meli_publishing.process_excel"

    data = fields.Binary(
        required=True,
        string="Worksheet to process"
    )
    pdata = fields.Binary(
        required=False,
        string="Worksheet processed",
    )
    name = fields.Char(
        'File Name',
    )
    state = fields.Selection(
        [('load', 'Load'),  # load spreadsheet
         ('download', 'Download'),  # download spreadsheet
         ('error', 'Error')],  # show errors
        default="load"
    )
    errors = fields.Html(
        default=_('<h1>We found the following errors</h1>'),
        readonly=True
    )

    @api.multi
    def add_error(self, tag, row=False, meli_code=False, sku=False,
                  default_code=False):
        for rec in self:
            if tag == 'not_found':
                rec.errors += _('<p>The product code %s from row %s on the '
                                'worksheet can not be found in the system.'
                                '</p>') % (meli_code, row)
            if tag == 'sku':
                rec.errors += _('<p>The sku from worksheet does not match with'
                                ' the product internal reference "%s" <> "%s" '
                                'at row %s in the worksheet.'
                                '</p>') % (sku, default_code, row)

    @api.multi
    def process_data(self, fp_name):
        PUB_CODE_COL = 2
        PRICE_COL = 7
        SKU_COL = 1
        STOCK_COL = 10
        FIRST_ROW = 4

        product_obj = self.env['product.product']

        # open worksheet
        wb = openpyxl.load_workbook(filename=fp_name,
                                    read_only=False,
                                    data_only=True)
        sheet = wb.active
        for reg in self:
            for row in range(FIRST_ROW, sheet.max_row):
                meli_code = sheet.cell(column=PUB_CODE_COL, row=row).value
                prod = product_obj.search([('meli_code', '=', meli_code)])
                if prod:
                    sheet.cell(column=SKU_COL,
                               row=row,
                               value=prod.default_code)
                    sheet.cell(column=PRICE_COL,
                               row=row,
                               value=prod.final_price)
                    sheet.cell(column=STOCK_COL,
                               row=row,
                               value=prod.virtual_available)
                else:
                    reg.state = 'error'
                    reg.add_error('not_found', row=row, meli_code=meli_code)

            if reg.state != 'error':
                wb.save(fp_name)

    @api.multi
    def load_file(self):
        for rec in self:

            # escribir la data en un archivo temporario
            data = base64.decodestring(rec.data)
            (fileno, fp_name) = tempfile.mkstemp('.xlsx', 'openerp_')
            with open(fp_name, "w") as worksheet:
                worksheet.write(data)

            # procesar el archivo
            self.process_data(fp_name)

            # leer el archivo temporario para hacer download
            with open(fp_name, "r") as worksheet:
                rec.pdata = base64.encodestring(worksheet.read())

            # cambiar la vista para que descarguen el archivo
            if rec.state != 'error':
                rec.state = 'download'
