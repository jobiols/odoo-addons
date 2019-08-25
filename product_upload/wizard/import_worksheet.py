# -*- coding: utf-8 -*-
# Part of Odoo. See LICENSE file for full copyright and licensing details.
from openerp import api, models, fields
import base64
import tempfile
import logging
from openerp.exceptions import UserError

HEADERS = [u'Referencia', u'Moneda', u'Costo', u'Precio', u'Descripción',
           u'IVA compras', u'IVA ventas', u'Código de barras',
           u'Código mercadolibre', u'parent']

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except (ImportError, IOError) as er:
    _logger.debug(er)

# info de como es la tabla, es una lista de tuplas y cada tupla tiene
# Nombre del campo, Requerido para Crear, Requerido para Actualizar y tipo
INFO = [
    {'name': 'default_code', 'create_req': True, 'update_req': True,
     'type': 'str'},
    {'name': 'currency', 'create_req': True, 'update_req': True,
     'type': 'currency'},
    {'name': 'cost', 'create_req': True, 'update_req': True, 'type': 'number'},
    {'name': 'price', 'create_req': True, 'update_req': True,
     'type': 'number'},
    {'name': 'name', 'create_req': True, 'update_req': False, 'type': 'str'},
    {'name': 'purchase_tax', 'create_req': True, 'update_req': False,
     'type': 'tax'},
    {'name': 'sale_tax', 'create_req': True, 'update_req': False,
     'type': 'tax'},
    {'name': 'barcode', 'create_req': False, 'update_req': False,
     'type': 'str'},
    {'name': 'meli', 'create_req': False, 'update_req': False, 'type': 'str'},
    {'name': 'parent', 'create_req': False, 'update_req': False,
     'type': 'default_code'}
]


class ImportWorksheet(models.TransientModel):
    _name = "product_upload.import_worksheet"

    data = fields.Binary(
        'File',
        required=True
    )
    name = fields.Char(
        'Filename',
        readonly=True
    )

    def check_validate(self, value, ok, exist, text, name):
        """     Value  Ok  Exist
                0      0   0       Error

                0      0   1       no salvar
                0      1   0       imposible
                0      1   1       imposible

                1      0   0       Error
                1      0   1       Error

                1      1   0       Salvar
                1      1   1       Salvar

                Salvar = Value & Ok
                Error = ~Value & ~Ok
                Error = ~Value & ~Exist
                No salvar = ~Value & Exist
        """
        if value and ok:
            return {name: value}

        if ((not value and not ok and not exist) or
                (value and not ok and not exist) or
                (value and not ok and exist)):
            self.add_error(text)
            return {}

        if not value and exist:
            return {}

    def read_data(self, sheet):
        """ Read the spreadsheet into a data structure
        """

        def check(info, col, row, rowix, exist):
            """ chequear los datos y generar errores si estan mal
            """
            if col > sheet.max_column:
                self.add_error(u'La variable col esta fuera de rango')

            # valor de la celda
            value = row[col].value
            _ = (rowix, HEADERS[col], sheet.title)
            err = u'Error en la fila %d columna "%s" de %s : ' % _
            name = info['name']

            # DEFAULT CODE ####################################################

            if col == 0:
                if isinstance(value, (str, unicode)):
                    value = value.strip()
                    return {name: value}
                else:
                    text = err + u'La celda debe contener el codigo de ' \
                                 u'producto. Pero se lee %s ' % value
                    self.add_error(text)
                return {}

            # CURRENCY ########################################################
            if col == 1:
                ok = (value == 'USD' or value == 'ARS')
                text = err + u'Moneda invalida, debe ser ARS o USD. ' \
                             u'Pero se lee %s ' % value
                return self.check_validate(value, ok, exist, text, name)

            # COST / PRICE ####################################################

            if col == 2 or col == 3:
                if isinstance(value, (float, int, long)):
                    return {name: value}
                else:
                    text = err + u'La celda debe contener un valor monetario' \
                                 u'. Pero se lee %s ' % value
                    self.add_error(text)
                return {}

            # DESCRIPCION #####################################################

            if col == 4:
                ok = isinstance(value, (str, unicode))
                text = err + u'La celda debe contener la descripcion del ' \
                             u'producto. Pero se lee %s ' % value
                return self.check_validate(value, ok, exist, text, name)

            # TAX #############################################################

            if col == 5 or col == 6:
                ok = isinstance(value, float) and (value < 1)
                text = err + u'La celda debe contener IVA, se espera ' \
                             u'un numero de punto flotante menor que ' \
                             u'uno, o un numero con formato de ' \
                             u'procentaje. Pero se lee %s ' % value
                return self.check_validate(value, ok, exist, text, name)

            # CODIGO DE BARRAS

            if col == 7:
                ok = isinstance(value, (str, unicode))
                if ok:
                    value = value.strip()
                text = err + u"La celda debe contener el codigo de barras " \
                             u"en formato de texto. Para lograr el formato " \
                             u"de texto escriba un apostrofo antes del " \
                             u"numero, de esta forma '454514815415. " \
                             u"Actualmente se lee %s " % value
                # ponemos exist = true porque no es un dato requerido.
                return self.check_validate(value, ok, True, text, name)

            # CODIGO DE MERCADOLIBRE

            if col == 8:
                ok = isinstance(value, (str, unicode))
                if ok:
                    value = value.strip()
                text = err + u"La celda debe contener el codigo de " \
                             u"mercadolibre en formato de texto. Sin " \
                             u"embargo se lee %s" % value
                # ponemos exist = true porque no es un dato requerido.
                return self.check_validate(value, ok, True, text, name)

            # PARENT

            if col == 9:
                if value and isinstance(value, (str, unicode)):
                    product_obj = self.env['product.template']
                    domain = [('default_code', '=', value)]
                    product = product_obj.search(domain)
                    if product:
                        return {name: value}
                    else:
                        text = err + u'La celda debe contener un codigo de ' \
                                     u'producto existente, el producto ' \
                                     u'definido en esta linea tomara el ' \
                                     u'precio de costo del producto de esta ' \
                                     u'columna.' \
                                     u'Sin embargo el producto %s que se ' \
                                     u'lee en la celda no existe.' % value
                        self.add_error(text)
                        return {}
                if value and not isinstance(value, (str, unicode)):
                    text = err + u'La celda debe contener texto. ' \
                                 u'Pero se lee %s ' % value
                    self.add_error(text)
                return {}

        if sheet.max_column < 10:
            self.add_error(u'Verifique los titulos en la planilla del '
                           u'proveedor %s parece que '
                           u'faltan columnas' % sheet.title)
            return

        ret = []
        # con rowix encuentro la row, no sirve hacer row[0].row porque si
        # esta vacia me da una excepcion.
        for rowix, row in enumerate(sheet.iter_rows(min_row=2, min_col=1,
                                                    max_row=sheet.max_row,
                                                    max_col=sheet.max_column)):
            line = {}

            exist = self.check_product_exist(row[0].value)
            if row[0].value is not None:
                for idx, info in enumerate(INFO):
                    value = check(info, idx, row, rowix + 2, exist)
                    line.update(value)

                # si no hay errores agrego la linea
                if not self.log.errors:
                    ret.append(line)
        return ret

    def check_product_exist(self, default_code):
        # si el default code no es texto no existe el producto
        if isinstance(default_code, (str, unicode)):
            default_code = default_code.strip()
        else:
            return False

        # chequear si el producto existe, con blancos stripeados
        prod_obj = self.env['product.template']
        prod = prod_obj.search([('default_code', '=', default_code)])
        return True if prod else False

    def process_data(self, data, vendor):
        """ Process data structure creating / updating products
        """

        def choose_tax(tax_sale):
            for tax in tax_sale:
                if tax.amount != 0:
                    # si no es cero es ese
                    return tax.id
                else:
                    # si es iva cero busco que sea exento
                    if tax.tax_group_id.afip_code == 2:
                        return tax.id

        product_obj = self.env['product.template']
        currency_obj = self.env['res.currency']
        barcode_obj = self.env['product.barcode']
        tax_obj = self.env['account.tax']

        for row in data:
            domain = [('default_code', '=', row['default_code'])]
            prod = product_obj.search(domain)
            data = {}

            if row.get('currency'):
                if self.env.user.currency_id.name != row['currency']:
                    pc = currency_obj.search([('name', '=', row['currency'])])
                    data['force_currency_id'] = pc.id

            if row.get('name'):
                data['name'] = row['name']
            if row.get('meli'):
                data['meli_code'] = row['meli']
            if row.get('parent'):
                data['parent_price_product'] = row['parent']

            data['type'] = 'product'
            data['invoice_policy'] = 'order'
            data['purchase_method'] = 'purchase'
            data['sale_delay'] = 0

            if not prod:
                data['default_code'] = row['default_code']
                prod = product_obj.create(data)
                self.add_create()
                _logger.info('create product %s' % prod.default_code)
            else:
                prod.ensure_one()
                prod.write(data)
                self.add_update()
                _logger.info('update product %s' % prod.default_code)

            if row.get('barcode'):
                barcode_obj.add_barcode(prod, row['barcode'])

            if row.get('purchase_tax'):
                # actualiza iva compras
                tax_purchase_id = tax_obj.search(
                    [('amount', '=', row['purchase_tax'] * 100),
                     ('tax_group_id.tax', '=', 'vat'),
                     ('type_tax_use', '=', 'purchase')])
                # analizando el iva
                tax = choose_tax(tax_purchase_id)

                # esto reemplaza todos los registros por el tax que es un id
                prod.supplier_taxes_id = [(6, 0, [tax])]

            if row.get('sale_tax'):
                # actualiza IVA ventas
                tax_sale_id = tax_obj.search(
                    [('amount', '=', row['sale_tax'] * 100),
                     ('tax_group_id.tax', '=', 'vat'),
                     ('type_tax_use', '=', 'sale')])
                # analizando el iva
                tax = choose_tax(tax_sale_id)

                # esto reemplaza todos los registros por el tax que es un id
                prod.taxes_id = [(6, 0, [tax])]

            prod.set_prices(row['cost'], vendor, price=row['price'],
                            manual=True)
            prod.set_invoice_cost()

    @api.multi
    def import_file(self):
        """ Se ejecuta desde el boton import file del wizard, sube la planilla
            y la deja en un archivo temporario
        """
        for rec in self:
            # obtener el record log
            log_obj = self.env['product_upload.log']
            current_id = self.env.context.get('default_active_id', False)
            self.log = log_obj.search([('id', '=', current_id)])
            if not self.log:
                # TODO esto no anda, ver porque
                raise Exception('unknown error')

            data = base64.decodestring(rec.data)
            (fileno, fp_name) = tempfile.mkstemp('.xlsx', 'openerp_')

            # escribir la planilla en un temporario
            with open(fp_name, 'w') as ws:
                ws.write(data)

            self.process_tmp_file(fp_name)

    def process_tmp_file(self, tmp_file):
        """ Procesa la planilla que esta en el archivo temporario
        """

        # leer la planilla del temporario en solo lectura
        try:
            wb = openpyxl.load_workbook(filename=tmp_file, read_only=True,
                                        data_only=True)
        except Exception, e:
            raise UserError(u'El archivo importado no es una planilla'
                            u' Microsoft Excel (.xlsx)')

        # cada hoja de la planilla es un vendor
        partner_obj = self.env['res.partner']
        vendors = wb.sheetnames
        for vendor in vendors:
            partner = partner_obj.search([('ref', '=', vendor)])
            if not partner:
                self.add_error(u'No se encuentra el '
                               u'proveedor con ref %s') % vendor
                break

            self.add_vendor(vendor)
            data = self.read_data(wb[vendor])

            if self.log.errors:
                break

            if data:
                self.process_data(data, vendor)
                self.log.state = 'done'

    def add_error(self, text):
        """ add an error to log
        """
        self.log.state = 'error'
        self.log.errors += 1
        self.log.error_ids.create({
            'name': text,
            'log_id': self.log.id
        })

    def add_create(self):
        """ Increment Created
        """
        self.log.created_products += 1

    def add_update(self):
        """ Increment Created
        """
        self.log.updated_products += 1

    def add_vendor(self, vendor):
        """ Add a vendor to a list
        """
        if not self.log.vendors:
            self.log.vendors = vendor
        else:
            self.log.vendors += ', ' + vendor
